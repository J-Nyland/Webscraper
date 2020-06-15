from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import time
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
import openpyxl
import numpy as np
import os
            import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

class GetEANCode():
    
    def __init__(self, url,
                    MyUserName =None,
                    password = None,
                    ChromeDriverPath = None,
                    ExcelDir = None,
                    fileName = None):
        self.url = url 
        self.MyUserName = MyUserName or  "martin.burrage@cmsdistribution.com"
        self.password = password or "Harrogate"
        self.ChromeDriverPath = ChromeDriverPath or r"C:\Program Files old driver\chromedriver_old.exe"
        self.ExcelDir = ExcelDir
        self.fileName = fileName
        self.excelPath = self.ExcelDir + '\\' + self.fileName + '.xlsx'
        print(self.excelPath)

    def GetEANCodeNow(self):
        #wb = getEAN.CompileWB(fileName)

        try:
            driver = webdriver.Chrome(self.ChromeDriverPath)
            
        except Exception:
 
body =  f""" <html>
                        <div class="container" style="background-color: #CC2026;">
                            <br> <br> 
                            <h1>The system failed to operate due to an error with the Chrome Driver. Please update the Chrome Driver being used</h1>
                    
                            <br>
                        </div>
 
                    </html>
                """
s = smtplib.SMTP('es12app2.corp.cmsdistribution.com:25')
msg = MIMEMultipart('alternative')
html  = MIMEText(body, 'html')
msg.attach(html)
msg['Subject'] = "Chrome Driver Error"   
msg['To'] = ", ".join("jonathan.nyland@cmsdistribution")
msg['html'] = body
 
s.sendmail("winshuttle@cms.com", "jonathan.nyland@cmsdistribution.com", msg.as_string())
        

        driver.get(self.url)

        attempt = 0
        while attempt < 10:
            try:
                time.sleep(5)
                emailTag = driver.find_element_by_id('login-email')
                emailTag.send_keys(self.MyUserName)
                break
            except Exception as e:
                print(e)
                attempt = attempt + 1
                pass

        time.sleep(1)

        passwordTag = driver.find_element_by_name('password')
        passwordTag.send_keys(self.password)

        time.sleep(1)
        signinButton = driver.find_element_by_class_name('buttonOrange').click()

        time.sleep(10)
        Select1 = Select(driver.find_element_by_class_name('numberTypeSelect'))
        Select1.select_by_visible_text('GTIN-13')

        time.sleep(5)

        CompanyPrefixNo = driver.find_element_by_class_name('chosen-search-input')
        time.sleep(2)
        CompanyPrefixNo.send_keys(5051868)
        time.sleep(2)
        CompanyPrefixNo.send_keys(Keys.RETURN)




        GoToNumberbankBtn = driver.find_element_by_class_name('buttonOrange').click()
                                            
        time.sleep(5)
        UseNextNumber = driver.find_element_by_xpath('/html/body/div[1]/div/div/div[2]/div[4]/div/div/div/div[5]/div/div/div[4]/a[1]').click()

        time.sleep(5)

      
        manCodeList = self.GetManCodes()
        for i, val  in enumerate(manCodeList):
            print(i)
            
            # this gets the EAN number xpath
            CopyNumber_text = driver.find_element_by_xpath('//*[@id="modal-body"]/div/div[1]/p').text
            
            #split element_text into [The word 'Number' and the actual number itself]
            split_Number = CopyNumber_text.split(' ') 
            EAN_num = split_Number[1]

            try:
                # convert to an integer to test whether the EANNum is a number (used for testing)
                int(EAN_num)

                #Product Name is apart of a list, had to use the index [0] to find the correct list element
                Product_Name = driver.find_elements_by_xpath('//*[@id="in-modal-name"]')[0]
                Product_Name.send_keys(val)

                # write EAN to excel
                self.WriteEanNumToExcel(i, EAN_num)
    
                time.sleep(5)
                #saveNextNumber =  driver.find_element_by_xpath('//*[@id="modal-body"]/div/div[5]/a[2]')
                #saveNextNumber.click() 
                #time.sleep(5)
                
            except:
                print(EAN_num)
                driver.quit()
                return False
        # final save
        Save_Button = driver.find_element_by_xpath('//*[@id="modal-body"]/div/div[5]/a[1]')
        Save_Button.click() 
        driver.quit()
                
        return True
            


        #Product Name is apart of a list, had to use the index [0] to find the correct list element
        # Product_Name = driver.find_elements_by_xpath('//*[@id="in-modal-name"]')[0]
        # Product_Name.send_keys("TEST")




        # Save_Button = driver.find_element_by_xpath('//*[@id="modal-body"]/div/div[5]/a[1]')
        # Save_Button.click()


        # UseNextNumber = driver.find_element_by_xpath(//div[@class='buttonOrange' and text()='Use next number'])
        # UseNextNumber.click()

        #//*[@id="content"]/div/div/div[5]/div/div/div[4]/a[1]

        #(//div[@class='buttonOrange' and text()='Use next number'])

    def WriteEanNumToExcel(self, rowNum, EANNum):
        try:
            wb = openpyxl.load_workbook(self.excelPath)
            wsForm = wb["Form"]
            # added the EAN to the excel
            print(rowNum)
            print(EANNum)
            wsForm[f"G{7+rowNum}"] = EANNum
            #wsForm.cell(row=5+rowNum, column=6).value = EAN_num            # Starting intiger is 0
            wb.save(self.excelPath)
            wb.Close(True)
        except:
            return False


    def GetManCodes(self):
        """This Def simply defines the wb that the data will be copied from."""

        try:
            
            wb = openpyxl.load_workbook(self.excelPath)
            wsForm = wb["Form"]
            list1 = [i[0].value for i in wsForm["F7":"F31"] if i[0].value != None]
            wb.close()
            return list1
        except Exception as e:
            raise Exception(e)

    def CompileWB(self, Dir, fileName):
        """This Def simply defines the wb that the data will be copied from."""
        try:
            wb = openpyxl.load_workbook(self.excelPath)
            return wb
        except Exception as e:
            raise Exception(e)

    def CompileWS(self, wb, sheetName):
        """This Def simply defines the ws that the data will be copied to."""

        try:
            ws = wb[sheetName]
            return ws
        except Exception as e:
            raise Exception(e)

    def WrieToExcel(self):
        wb = openpyxl.load_workbook(self.excelPath)
        wsForm = wb["Form"]
        wsForm[f"G9"] = "this is a test"
        wb.save(self.excelPath)
        return


url = "https://mygs1.gs1uk.org/My-GS1/checkout.ssp?is=login&login=T&origin=customercenter#login-register"
filename = r"BOM Winshuttle Test"
ExcelDir = r'C:\Users\nylandjon\Documents\Projects\Webscraper'

getEAN = GetEANCode(url, fileName=filename, ExcelDir=ExcelDir)
EANNum = getEAN.GetEANCodeNow()


if EANNum:
    print("Success ")
else:
    print("failure") 


# Dir = r'full path excluding filename'
# wb = getEAN.CompileWB(Dir, 'fileName')
# ws = self.CompileWS(wb, 'sheetName')
# getCellValue = ws.cell(row=int, column=int).value                # Starting intiger is 0
# ws.cell(row=int, column=int).value = "WriteCellValue"            # Starting intiger is 0
# wb.Close(True)

# MyUserName = "martin.burrage@cmsdistribution.com"
# MyPassword = "Harrogate"

# Path = "C:\Program Files JN\chromedriver.exe"
# driver = webdriver.Chrome(Path)

# driver.get("https://mygs1.gs1uk.org/My-GS1/checkout.ssp?is=login&login=T&origin=customercenter#login-register")

# attempt = 0
# while attempt < 10:
#     try:
#         time.sleep(5)
#         emailTag = driver.find_element_by_id('login-email')
#         emailTag.send_keys(MyUserName)
#         break
#     except Exception as e:
#         print(e)
#         attempt = attempt + 1
#         pass

# time.sleep(1)

# passwordTag = driver.find_element_by_name('password')
# passwordTag.send_keys(MyPassword)

# time.sleep(1)
# signinButton = driver.find_element_by_class_name('buttonOrange').click()

# time.sleep(10)
# Select1 = Select(driver.find_element_by_class_name('numberTypeSelect'))
# Select1.select_by_visible_text('GTIN-13')

# time.sleep(5)

# CompanyPrefixNo = driver.find_element_by_class_name('chosen-search-input')
# time.sleep(2)
# CompanyPrefixNo.send_keys(5051868)
# time.sleep(2)
# CompanyPrefixNo.send_keys(Keys.RETURN)

# GoToNumberbankBtn = driver.find_element_by_class_name('buttonOrange').click()
                                    
# time.sleep(5)
# UseNextNumber = driver.find_element_by_xpath('/html/body/div[1]/div/div/div[2]/div[4]/div/div/div/div[5]/div/div/div[4]/a[1]').click()

# time.sleep(5)
# CopyNumber_text = driver.find_element_by_xpath('//*[@id="modal-body"]/div/div[1]/p').text
# print(CopyNumber_text)
# #split element_text into [The word 'Number' and the actual number itself]
# split_Number = CopyNumber_text.split(' ') 
# EAN_num = split_Number[1]
# print(EAN_num)

# #Product Name is apart of a list, had to use the index [0] to find the correct list element
# Product_Name = driver.find_elements_by_xpath('//*[@id="in-modal-name"]')[0]
# Product_Name.send_keys("TEST")




# Save_Button = driver.find_element_by_xpath('//*[@id="modal-body"]/div/div[5]/a[1]')
# Save_Button.click()


# # UseNextNumber = driver.find_element_by_xpath(//div[@class='buttonOrange' and text()='Use next number'])
# # UseNextNumber.click()

# #//*[@id="content"]/div/div/div[5]/div/div/div[4]/a[1]

# #(//div[@class='buttonOrange' and text()='Use next number'])






